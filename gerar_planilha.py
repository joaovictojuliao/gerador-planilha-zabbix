from pyzabbix import ZabbixAPI
import pandas as pd
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

def gerar_excel():
    ZABBIX_SERVER = "https://pulse.jelastic.saveincloud.net"
    USERNAME = "hoster"
    PASSWORD = "3aWBa103nfiF86TPdkuo"
    GROUP_NAME = "HARDNODES"

    time_till = int(time.time())
    time_from = time_till - (7 * 24 * 60 * 60)

    zapi = ZabbixAPI(ZABBIX_SERVER)
    zapi.session.verify = False  # ← desativa verificação de SSL
    zapi.login(USERNAME, PASSWORD)

    grupo = zapi.hostgroup.get(filter={"name": GROUP_NAME})
    if not grupo:
        return None

    groupid = grupo[0]['groupid']

    hosts = zapi.host.get(
        output=["hostid", "host"],
        groupids=groupid,
        filter={"status": "0"}
    )
    if not hosts:
        return None

    def get_trend_value(hostid, key, mode="avg"):
        item = zapi.item.get(hostids=hostid, search={"name": key}, output=["itemid"])
        if not item:
            return None
        itemid = item[0]['itemid']
        try:
            trends = zapi.trend.get(itemids=itemid, time_from=time_from, time_till=time_till)
            if trends:
                if mode == "max":
                    valores = [float(t['value_max']) for t in trends]
                    return max(valores) if valores else None
                else:
                    valores = [float(t['value_avg']) for t in trends]
                    return sum(valores) / len(valores) if valores else None
        except:
            return None
        return None

    def get_disk_free(hostid):
        pontos = ["/vz", "/"]
        for ponto in pontos:
            item = zapi.item.get(
                hostids=hostid,
                search={"key_": f"vfs.fs.size[{ponto},free]"},
                output=["itemid"]
            )
            if item:
                itemid = item[0]['itemid']
                try:
                    history = zapi.history.get(
                        itemids=itemid,
                        history=3,
                        sortfield="clock",
                        sortorder="DESC",
                        limit=1
                    )
                    if history:
                        valor_bytes = float(history[0]['value'])
                        valor_gb = round(valor_bytes / 1024 / 1024 / 1024, 2)
                        return valor_gb
                except:
                    return None
        return None

    def identificar_tipo(hostname):
        if any(p in hostname for p in ["vinw", "vin-w"]):
            return "Windows"
        elif any(p in hostname for p in ["vinl", "vin-l"]):
            return "Linux"
        elif "bsb" in hostname:
            return "Brasilia"
        elif "bkp" in hostname:
            return "Backup"
        elif any(p in hostname for p in ["hn05.nordeste", "hn07.ce"]):
            return "Fortaleza Windows"
        else:
            return "Fortaleza"

    data = []
    for host in hosts:
        hostname = host['host']
        load15_avg = get_trend_value(host['hostid'], "Processor load15", mode="avg")
        load15_max = get_trend_value(host['hostid'], "Processor load15", mode="max")
        disk_free = get_disk_free(host['hostid'])

        data.append({
            "Host": hostname,
            "Tipo": identificar_tipo(hostname),
            "Load15 AVG (Média 7 dias)": int(load15_avg) if load15_avg is not None else "N/A",
            "Load15 MAX (Máximo 7 dias)": int(load15_max) if load15_max is not None else "N/A",
            "Disco Livre (GB)": int(disk_free) if disk_free is not None else "N/A"
        })

    df = pd.DataFrame(data)

    df["Host"] = df["Host"].str.replace(r"^hm019-vinl$", "hn019-vinL", case=False, regex=True)

    ordem_tipo = {
        "Windows": 1,
        "Linux": 2,
        "Fortaleza Windows": 3,
        "Fortaleza": 4,
        "Brasilia": 5,
        "Backup": 6,
        "Outro": 7
    }
    df = df.sort_values(
        by=["Tipo", "Host"],
        key=lambda x: x.map(ordem_tipo) if x.name == "Tipo" else x
    )

    base_dir = os.path.dirname(os.path.abspath(__file__))
    pasta_saida = os.path.join(base_dir, "tmp_planilhas")
    os.makedirs(pasta_saida, exist_ok=True)  # Cria se não existir

    data_atual = datetime.now().strftime('%d-%m-%y_%H')
    arquivo_saida = os.path.join(pasta_saida, f"LoadDisk {data_atual}.xlsx")
    df.to_excel(arquivo_saida, index=False)

    wb = load_workbook(arquivo_saida)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    wb.save(arquivo_saida)
    return arquivo_saida


